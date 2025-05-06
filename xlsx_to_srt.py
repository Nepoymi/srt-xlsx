"""
Перед выполнением:
Установить python (Для windows - https://www.python.org/downloads/)
Установить pip (Для windows - в cmd выполнить "python -m ensurepip --default-pip"
Установить модуль openpyxl (pip install openpyxl)

Для Linux - дополнительно читать README.md
"""
from openpyxl import load_workbook
from tkinter import Tk, filedialog

def convert_xlsx_to_srt():
    root = Tk()
    root.withdraw()
    
    # Выбор исходного файла
    input_xlsx = filedialog.askopenfilename(
        title="Выберите XLSX файл",
        filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
    )
    if not input_xlsx:
        return
    
    # Выбор выходного файла
    output_srt = filedialog.asksaveasfilename(
        title="Сохранить как SRT",
        defaultextension=".srt",
        filetypes=[("SRT файлы", "*.srt"), ("Все файлы", "*.*")]
    )
    if not output_srt:
        return
    
    wb = load_workbook(input_xlsx)
    ws = wb.active
    
    with open(output_srt, 'w', encoding='utf-8') as f:
        for row in ws.iter_rows(min_row=2, values_only=True):
            index, timestamp, subtitles = row
            
            if not index or not timestamp:
                continue
                
            subs = str(subtitles).split('\n') if subtitles else ['']
            
            f.write(f"{index}\n")
            f.write(f"{timestamp}\n")
            f.write('\n'.join(subs) + "\n\n")

if __name__ == "__main__":
    convert_xlsx_to_srt()
